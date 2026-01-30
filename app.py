#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Calendario Istruttori - Flask App
"""

from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from datetime import datetime, timedelta
import calendar as cal_module
from database import (get_db, calcola_data_fine, get_festivi_italiani, get_festivi_completi, 
                     verifica_sovrapposizione, log_action, init_db, hash_password, verify_password,
                     get_utente_by_username, get_utente_by_id, crea_utente, lista_utenti, 
                     aggiorna_ultimo_accesso, registra_audit, get_audit_log)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from functools import wraps
import os
import urllib.request
from pathlib import Path

app = Flask(__name__)
app.config['JSON_AS_ASCII'] = False
app.secret_key = os.environ.get('SECRET_KEY', 'dev-key-change-in-production-' + os.urandom(12).hex())

# Ensure static libraries are downloaded
def ensure_static_libs():
    """Ensure Bootstrap and FontAwesome files are available"""
    lib_dir = Path(__file__).parent / 'static' / 'lib'
    
    files_to_check = {
        'bootstrap/bootstrap.min.css': 'https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css',
        'bootstrap/bootstrap.bundle.min.js': 'https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js',
        'fontawesome/all.min.css': 'https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css',
        'fontawesome/webfonts/fa-solid-900.woff2': 'https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/webfonts/fa-solid-900.woff2',
        'fontawesome/webfonts/fa-solid-900.ttf': 'https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/webfonts/fa-solid-900.ttf',
    }
    
    # Migrazione: se i font sono nella vecchia posizione, spostali in webfonts/
    legacy_woff2 = lib_dir / 'fontawesome' / 'fa-solid-900.woff2'
    legacy_ttf = lib_dir / 'fontawesome' / 'fa-solid-900.ttf'
    webfonts_dir = lib_dir / 'fontawesome' / 'webfonts'
    if legacy_woff2.exists() and not (webfonts_dir / 'fa-solid-900.woff2').exists():
        try:
            webfonts_dir.mkdir(parents=True, exist_ok=True)
            legacy_woff2.replace(webfonts_dir / 'fa-solid-900.woff2')
        except Exception as e:
            print(f"⚠️ Could not migrate woff2: {e}")
    if legacy_ttf.exists() and not (webfonts_dir / 'fa-solid-900.ttf').exists():
        try:
            webfonts_dir.mkdir(parents=True, exist_ok=True)
            legacy_ttf.replace(webfonts_dir / 'fa-solid-900.ttf')
        except Exception as e:
            print(f"⚠️ Could not migrate ttf: {e}")

    for rel_path, url in files_to_check.items():
        file_path = lib_dir / rel_path
        if not file_path.exists():
            try:
                file_path.parent.mkdir(parents=True, exist_ok=True)
                print(f"📥 Downloading {rel_path}...")
                urllib.request.urlretrieve(url, str(file_path))
                print(f"✅ Downloaded {rel_path}")
            except Exception as e:
                print(f"⚠️ Could not download {rel_path}: {e}")

# Try to ensure libs on startup (non-blocking)
try:
    ensure_static_libs()
except Exception as e:
    print(f"⚠️ Static libs setup error: {e}")

# Initialize DB on startup (ensures tables exist on Render)
try:
    init_db()
except Exception as e:
    print(f"⚠️ init_db() error: {e}")

# Ensure requested admin user exists (idempotent)
try:
    existing_user = get_utente_by_username('Pasturenzi')
    if not existing_user:
        crea_utente('Pasturenzi', '', 'Pasturenzi', '', 'PaolinoCZ', 'Admin')
        print("✅ Utente admin 'Pasturenzi' creato")
except Exception as e:
    print(f"⚠️ Errore creazione utente Pasturenzi: {e}")


# ==================== DECORATORI PROTEZIONE ====================

def require_login(f):
    """Decoratore: richiede login"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Controlla se database auth è inizializzato
        try:
            conn = get_db()
            utenti_count = conn.execute("SELECT COUNT(*) as cnt FROM utenti").fetchone()['cnt']
            conn.close()
            
            if utenti_count == 0:
                # Database non inizializzato, redirect a setup
                return redirect(url_for('setup_admin'))
        except Exception as e:
            # Errore lettura database - probabilmente tabelle non esistono
            # Redirect a setup per inizializzare
            print(f"⚠️ Database check error: {e}")
            return redirect(url_for('setup_admin'))
        
        # Database ok, controlla sessione
        if 'utente_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def require_role(role_name):
    """Decoratore: richiede ruolo specifico"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'utente_id' not in session:
                return redirect(url_for('login'))
            
            utente = get_utente_by_id(session['utente_id'])
            if not utente or utente['ruolo_nome'] != role_name:
                return render_template('errore.html', 
                    messaggio='Non hai permesso di accedere a questa pagina'), 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def require_editor(f):
    """Decoratore: richiede Editor o Admin"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'utente_id' not in session:
            return redirect(url_for('login'))
        
        utente = get_utente_by_id(session['utente_id'])
        if not utente or utente['ruolo_nome'] not in ['Admin', 'Editor']:
            return render_template('errore.html',
                messaggio='Solo Editor e Admin possono modificare'), 403
        return f(*args, **kwargs)
    return decorated_function


@app.template_filter('itdate')
def itdate(value):
    """Formatta 'YYYY-MM-DD' in 'DD/MM/YYYY' (fallback su parsing date)."""
    try:
        if not value:
            return ''
        s = str(value)
        parts = s.split('T')[0].split('-')
        if len(parts) == 3:
            yyyy, mm, dd = parts
            return f"{dd.zfill(2)}/{mm.zfill(2)}/{yyyy}"
        d = datetime.strptime(s, '%Y-%m-%d')
        return f"{d.day:02d}/{d.month:02d}/{d.year}"
    except Exception:
        try:
            d = datetime.fromisoformat(str(value).split('T')[0])
            return f"{d.day:02d}/{d.month:02d}/{d.year}"
        except Exception:
            return str(value)


def giorni_lavorativi_tra(data_inizio, data_fine):
    """Calcola giorni lavorativi (lunedì-venerdì) tra due date"""
    if not data_inizio or not data_fine:
        return 0
    try:
        start = datetime.strptime(data_inizio, '%Y-%m-%d').date()
        end = datetime.strptime(data_fine, '%Y-%m-%d').date()
    except:
        return 0
    if end < start:
        return 0
    giorni = 0
    current = start
    while current <= end:
        if current.weekday() < 5:
            giorni += 1
        current += timedelta(days=1)
    return giorni

def aggiungi_giorni_lavorativi(data_inizio, giorni_da_aggiungere):
    """Restituisce una nuova data aggiungendo N giorni lavorativi (lun-ven, esclusi festivi+custom)."""
    if giorni_da_aggiungere <= 0:
        return data_inizio
    try:
        corrente = datetime.strptime(data_inizio, '%Y-%m-%d')
    except Exception:
        return data_inizio
    anno = corrente.year
    festivi = set(get_festivi_completi(anno))

    aggiunti = 0
    while aggiunti < giorni_da_aggiungere:
        corrente += timedelta(days=1)
        # Aggiorna festivi se cambio anno
        if corrente.year != anno:
            anno = corrente.year
            festivi = set(get_festivi_completi(anno))
        data_str = corrente.strftime('%Y-%m-%d')
        is_weekend = corrente.weekday() >= 5
        is_festivo = data_str in festivi
        if not is_weekend and not is_festivo:
            aggiunti += 1
    return corrente.strftime('%Y-%m-%d')

# Nomi mesi in italiano
MESI_ITALIANI = [
    '', 'Gennaio', 'Febbraio', 'Marzo', 'Aprile', 'Maggio', 'Giugno',
    'Luglio', 'Agosto', 'Settembre', 'Ottobre', 'Novembre', 'Dicembre'
]

# ============================================================================
# ROUTES AUTENTICAZIONE
# ============================================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Pagina login"""
    # Auto-crea admin se non esistono utenti
    conn = get_db()
    count = conn.execute('SELECT COUNT(*) as cnt FROM utenti').fetchone()
    if count['cnt'] == 0:
        # Crea utente admin automaticamente
        from database import crea_utente
        crea_utente('3102011', '3102011@trenord.it', 'Admin', 'Calendario', 'Qaqqa1234.', 'Admin')
    conn.close()
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        utente = get_utente_by_username(username)
        if utente and verify_password(password, utente['password_hash']):
            session['utente_id'] = utente['id']
            session['username'] = utente['username']
            session['ruolo'] = utente['ruolo_nome']
            aggiorna_ultimo_accesso(utente['id'])
            registra_audit(utente['id'], 'LOGIN', ip_address=request.remote_addr)
            return redirect(url_for('index'))
        else:
            return render_template('login.html', errore='Username o password errati')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    """Logout"""
    if 'utente_id' in session:
        registra_audit(session.get('utente_id'), 'LOGOUT', ip_address=request.remote_addr)
    session.clear()
    return redirect(url_for('login'))

@app.route('/setup-admin')
def setup_admin():
    """Setup iniziale - crea database e admin - NON PROTETTO"""
    try:
        # Drop tabelle di auth vecchie (potrebbero avere struttura diversa)
        conn = get_db()
        c = conn.cursor()
        
        for table in ['permessi_utente', 'utenti', 'audit_log', 'ruoli']:
            try:
                c.execute(f'DROP TABLE IF EXISTS {table}')
            except:
                pass
        
        conn.commit()
        conn.close()
        
        # Ricrea tabelle con init_db
        from database import init_db
        init_db(drop_auth_tables=True)
        
        # Connetti di nuovo e inserisci dati
        conn = get_db()
        
        # Ruoli
        ruoli_default = [
            ("Admin", "Accesso totale, gestione utenti e permessi"),
            ("Editor", "Può modificare impegni, corsi, attività"),
            ("Viewer", "Accesso in sola lettura"),
        ]
        
        for nome, descrizione in ruoli_default:
            conn.execute('INSERT OR IGNORE INTO ruoli (nome, descrizione) VALUES (?, ?)', 
                        (nome, descrizione))
        conn.commit()
        
        # Prendi ID Admin
        ruolo_admin = conn.execute("SELECT id FROM ruoli WHERE nome = 'Admin'").fetchone()
        if not ruolo_admin:
            raise Exception("Impossibile trovare ruolo Admin")
        
        ruolo_id = ruolo_admin['id']
        
        # Crea admin se non esiste
        existing = conn.execute("SELECT * FROM utenti WHERE username = '3102011'").fetchone()
        if not existing:
            conn.execute('''
                INSERT INTO utenti (username, email, nome, cognome, password_hash, ruolo_id, attivo)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', ('3102011', '3102011@trenord.it', 'Admin', 'Calendario', 'Qaqqa1234.', ruolo_id, 1))
            conn.commit()
        
        conn.close()
        
        return '''
        <html>
        <body style="font-family: Arial; text-align: center; padding: 50px;">
            <h1 style="color: green;">✅ Setup Completato!</h1>
            <p>Database inizializzato con successo</p>
            <p><strong>Username:</strong> 3102011</p>
            <p><strong>Password:</strong> Qaqqa1234.</p>
            <br>
            <a href="/login" style="background: #667eea; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">
                Vai al Login →
            </a>
        </body>
        </html>
        '''
    except Exception as e:
        import traceback
        return f'''
        <html>
        <body style="font-family: Arial; padding: 50px;">
            <h1 style="color: red;">❌ Errore Setup</h1>
            <p><strong>Errore:</strong> {str(e)}</p>
            <pre style="background: #eee; padding: 10px; overflow-x: auto; font-size: 12px;">{traceback.format_exc()}</pre>
            <br>
            <a href="/setup-admin" style="background: #ffc107; color: black; padding: 10px 20px; text-decoration: none; border-radius: 5px;">
                Riprova Setup
            </a>
        </body>
        </html>
        ''', 500

@app.route('/debug-db')
def debug_db():
    """Diagnostica database"""
    try:
        import os
        conn = get_db()
        
        # Lista tabelle
        tabelle = conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
        tabelle_list = [t['name'] for t in tabelle]
        
        # Conta utenti
        try:
            utenti_count = conn.execute("SELECT COUNT(*) as cnt FROM utenti").fetchone()
            utenti_count = utenti_count['cnt'] if utenti_count else 0
        except:
            utenti_count = "Tabella non esiste"
        
        # Conta ruoli
        try:
            ruoli_count = conn.execute("SELECT COUNT(*) as cnt FROM ruoli").fetchone()
            ruoli_count = ruoli_count['cnt'] if ruoli_count else 0
        except:
            ruoli_count = "Tabella non esiste"
        
        # Prova a recuperare utente admin
        try:
            admin_user = conn.execute("SELECT * FROM utenti WHERE username = '3102011'").fetchone()
            admin_info = dict(admin_user) if admin_user else "Non trovato"
        except Exception as e:
            admin_info = f"Errore: {e}"
        
        conn.close()
        
        # DB path
        db_path = os.path.abspath('calendario.db')
        db_exists = os.path.exists(db_path)
        
        return f'''
        <html>
        <head><style>
            body {{ font-family: monospace; padding: 20px; background: #f5f5f5; }}
            .box {{ background: white; padding: 20px; margin: 10px 0; border-radius: 5px; }}
            .ok {{ color: green; }}
            .error {{ color: red; }}
            pre {{ background: #eee; padding: 10px; overflow-x: auto; }}
        </style></head>
        <body>
            <h1>🔍 Diagnostica Database</h1>
            
            <div class="box">
                <h3>Database File</h3>
                <p><strong>Path:</strong> {db_path}</p>
                <p><strong>Esiste:</strong> <span class="{'ok' if db_exists else 'error'}">{db_exists}</span></p>
            </div>
            
            <div class="box">
                <h3>Tabelle Presenti</h3>
                <pre>{tabelle_list}</pre>
            </div>
            
            <div class="box">
                <h3>Conteggi</h3>
                <p><strong>Utenti:</strong> {utenti_count}</p>
                <p><strong>Ruoli:</strong> {ruoli_count}</p>
            </div>
            
            <div class="box">
                <h3>Utente Admin (3102011)</h3>
                <pre>{admin_info}</pre>
            </div>
            
            <br>
            <a href="/setup-admin" style="background: #667eea; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">
                Forza Setup
            </a>
            <a href="/login" style="background: #28a745; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px; margin-left: 10px;">
                Prova Login
            </a>
        </body>
        </html>
        '''
    except Exception as e:
        return f'''
        <html>
        <body style="font-family: Arial; padding: 50px;">
            <h1 style="color: red;">❌ Errore Debug</h1>
            <pre>{str(e)}</pre>
        </body>
        </html>
        ''', 500

# ============================================================================
# ROUTES PAGINE
# ============================================================================

@app.route('/')
@require_login
def index():
    """Home page"""
    return render_template('index.html', utente={'username': session.get('username'), 'ruolo': session.get('ruolo')})

@app.route('/impegni')
@require_login
def impegni():
    """Pagina gestione assenze istruttori"""
    conn = get_db()
    
    # Lista istruttori
    istruttori = conn.execute('SELECT * FROM istruttori WHERE attivo = 1 ORDER BY nome').fetchall()
    
    # Lista attività - SOLO categoria ASSENZA
    attivita = conn.execute('''
        SELECT * FROM tipi_attivita 
        WHERE categoria = 'ASSENZA'
        ORDER BY nome
    ''').fetchall()
    
    # Lista impegni con join - SOLO categoria ASSENZA
    impegni_list = conn.execute('''
        SELECT i.*, ist.nome as istruttore_nome, ta.nome as attivita_nome, ta.colore
        FROM impegni i
        JOIN istruttori ist ON i.istruttore_id = ist.id
        JOIN tipi_attivita ta ON i.attivita_id = ta.id
        WHERE ta.categoria = 'ASSENZA'
        ORDER BY i.data_inizio DESC
    ''').fetchall()
    
    conn.close()
    
    return render_template('impegni.html', 
                         istruttori=istruttori, 
                         attivita=attivita,
                         impegni=impegni_list)

@app.route('/calendario/<int:anno>')
def calendario(anno):
    """Pagina calendario per anno specifico"""
    conn = get_db()
    
    # Lista istruttori
    istruttori = conn.execute('SELECT * FROM istruttori WHERE attivo = 1 ORDER BY nome').fetchall()
    
    # Impegni dell'anno
    impegni_anno_raw = conn.execute('''
        SELECT i.*, ist.nome as istruttore_nome, ta.nome as attivita_nome, ta.colore
        FROM impegni i
        JOIN istruttori ist ON i.istruttore_id = ist.id
        JOIN tipi_attivita ta ON i.attivita_id = ta.id
        WHERE 
            (strftime('%Y', i.data_inizio) = ? OR strftime('%Y', i.data_fine) = ?)
            OR (i.data_inizio <= ? AND i.data_fine >= ?)
    ''', (str(anno), str(anno), f"{anno}-12-31", f"{anno}-01-01")).fetchall()
    
    # Converti Row in dict per JSON serialization
    impegni_anno = [dict(row) for row in impegni_anno_raw]
    
    # Sostituzioni dell'anno
    sostituzioni_raw = conn.execute('''
        SELECT s.*, 
               io.nome as istruttore_originale, 
               isost.nome as istruttore_sostituto,
               i.attivita_id, ta.colore
        FROM sostituzioni s
        JOIN istruttori io ON s.istruttore_originale_id = io.id
        JOIN istruttori isost ON s.istruttore_sostituto_id = isost.id
        JOIN impegni i ON s.impegno_id = i.id
        JOIN tipi_attivita ta ON i.attivita_id = ta.id
        WHERE strftime('%Y', s.data_sostituzione) = ?
    ''', (str(anno),)).fetchall()
    
    sostituzioni_anno = [dict(row) for row in sostituzioni_raw]
    
    conn.close()
    
    # Calcola giorni dell'anno
    num_giorni = 366 if cal_module.isleap(anno) else 365
    
    # Festivi italiani
    festivi = get_festivi_italiani(anno)
    
    # Crea array giorno_info: per ogni giorno dell'anno, salva weekday (0=Dom, 6=Sab)
    giorno_info = []
    from datetime import date
    for g in range(1, num_giorni + 1):
        data = date(anno, 1, 1) + timedelta(days=g-1)
        weekday = data.weekday()  # Python: 0=Lun, 6=Dom
        # Converti a formato JS: 0=Dom, 1=Lun, ..., 6=Sab
        weekday_js = (weekday + 1) % 7
        giorno_info.append({
            'giorno': g,
            'weekday': weekday_js,
            'data': data.strftime('%Y-%m-%d')
        })
    
    # Crea struttura mesi con nomi italiani
    mesi = []
    for mese in range(1, 13):
        giorni_mese = cal_module.monthrange(anno, mese)[1]
        mesi.append({
            'numero': mese,
            'nome': MESI_ITALIANI[mese],
            'giorni': giorni_mese
        })
    
    return render_template('calendario.html',
                         anno=anno,
                         num_giorni=num_giorni,
                         mesi=mesi,
                         istruttori=istruttori,
                         impegni=impegni_anno,
                         sostituzioni=sostituzioni_anno,
                         festivi=festivi,
                         giorno_info=giorno_info)

@app.route('/corso/<path:id_corso>')
def corso(id_corso):
    """Pagina dettaglio corso con calendario"""
    conn = get_db()
    
    # Impegni del corso
    impegni_corso_raw = conn.execute('''
        SELECT i.*, ist.nome as istruttore_nome, ta.nome as attivita_nome, ta.colore
        FROM impegni i
        JOIN istruttori ist ON i.istruttore_id = ist.id
        JOIN tipi_attivita ta ON i.attivita_id = ta.id
        WHERE i.id_corso = ?
        ORDER BY i.data_inizio
    ''', (id_corso,)).fetchall()
    
    # Converti Row in dict per template
    impegni_corso = [dict(row) for row in impegni_corso_raw]
    
    if not impegni_corso:
        conn.close()
        return "Corso non trovato", 404
    
    # Recupera istruttore di riferimento (se presente)
    istruttore_riferimento_id = impegni_corso[0].get('istruttore_riferimento') if impegni_corso else None
    
    # Recupera info logistiche dal primo impegno
    luogo_corso = impegni_corso[0].get('luogo') if impegni_corso else None
    aula_corso = impegni_corso[0].get('aula') if impegni_corso else None
    posti_corso = impegni_corso[0].get('posti') if impegni_corso else None
    
    # Lista istruttori UNICI per il calendario (senza duplicati)
    istruttori_unici = []
    istruttori_visti = set()
    for imp in impegni_corso:
        if imp['istruttore_id'] not in istruttori_visti:
            istruttori_unici.append({
                'id': imp['istruttore_id'], 
                'nome': imp['istruttore_nome'],
                'is_riferimento': (imp['istruttore_id'] == istruttore_riferimento_id)
            })
            istruttori_visti.add(imp['istruttore_id'])
    
    # Trova periodo totale del corso (min data_inizio, max data_fine)
    data_inizio_corso = min(imp['data_inizio'] for imp in impegni_corso)
    data_fine_corso = max(imp['data_fine'] for imp in impegni_corso)
    
    # IDs degli istruttori coinvolti nel corso
    istruttori_ids = list(istruttori_visti)
    
    # Recupera assenze (impegni senza id_corso) degli istruttori del corso che ricadono nel periodo
    if istruttori_ids:
        placeholders = ','.join('?' * len(istruttori_ids))
        assenze_raw = conn.execute(f'''
            SELECT i.*, ist.nome as istruttore_nome, ta.nome as attivita_nome, ta.colore
            FROM impegni i
            JOIN istruttori ist ON i.istruttore_id = ist.id
            JOIN tipi_attivita ta ON i.attivita_id = ta.id
            WHERE i.istruttore_id IN ({placeholders})
                AND i.id_corso IS NULL
                AND (i.data_inizio <= ? AND i.data_fine >= ?)
            ORDER BY i.data_inizio
        ''', (*istruttori_ids, data_fine_corso, data_inizio_corso)).fetchall()
        
        assenze = [dict(row) for row in assenze_raw]
        # Aggiungi assenze agli impegni del corso per visualizzazione calendario
        impegni_corso.extend(assenze)
    
    # Calcola numero di giorni
    from datetime import datetime, timedelta
    inizio = datetime.strptime(data_inizio_corso, '%Y-%m-%d')
    fine = datetime.strptime(data_fine_corso, '%Y-%m-%d')
    num_giorni = (fine - inizio).days + 1
    
    # Crea lista di tutti i giorni
    giorni = []
    for i in range(num_giorni):
        giorno = inizio + timedelta(days=i)
        giorni.append({
            'data': giorno.strftime('%Y-%m-%d'),
            'giorno': giorno.day,
            'mese': MESI_ITALIANI[giorno.month],
            'giorno_settimana': ['Lun', 'Mar', 'Mer', 'Gio', 'Ven', 'Sab', 'Dom'][giorno.weekday()],
            'indice': i + 1
        })
    
    # Sostituzioni del corso
    sostituzioni_corso_raw = conn.execute('''
        SELECT s.*, 
               io.nome as istruttore_originale, 
               isost.nome as istruttore_sostituto,
               i.attivita_id, ta.colore, ta.nome as attivita_nome
        FROM sostituzioni s
        JOIN istruttori io ON s.istruttore_originale_id = io.id
        JOIN istruttori isost ON s.istruttore_sostituto_id = isost.id
        JOIN impegni i ON s.impegno_id = i.id
        JOIN tipi_attivita ta ON i.attivita_id = ta.id
        WHERE i.id_corso = ?
    ''', (id_corso,)).fetchall()
    
    sostituzioni_corso = [dict(row) for row in sostituzioni_corso_raw]
    
    conn.close()
    
    return render_template('corso.html',
                         id_corso=id_corso,
                         impegni=impegni_corso,
                         istruttori_unici=istruttori_unici,
                         istruttore_riferimento_id=istruttore_riferimento_id,
                         luogo=luogo_corso,
                         aula=aula_corso,
                         posti=posti_corso,
                         giorni=giorni,
                         num_giorni=num_giorni,
                         data_inizio=data_inizio_corso,
                         data_fine=data_fine_corso,
                         sostituzioni=sostituzioni_corso)

@app.route('/vista_istruttori')
def vista_istruttori():
    """Pagina per selezionare un istruttore e vedere statistiche"""
    conn = get_db()
    
    # Lista istruttori attivi con statistiche anno corrente
    anno_corrente = 2025
    
    istruttori_raw = conn.execute('''
        SELECT i.id, i.nome, i.attivo,
               COUNT(DISTINCT imp.id) as num_impegni,
               COALESCE(SUM(imp.giorni_lavorativi), 0) as giorni_lavorativi,
               MIN(imp.data_inizio) as prima_attivita,
               MAX(imp.data_fine) as ultima_attivita
        FROM istruttori i
        LEFT JOIN impegni imp ON i.id = imp.istruttore_id 
            AND (strftime('%Y', imp.data_inizio) = ? OR strftime('%Y', imp.data_fine) = ?)
        WHERE i.attivo = 1
        GROUP BY i.id, i.nome, i.attivo
        ORDER BY i.nome
    ''', (str(anno_corrente), str(anno_corrente))).fetchall()
    
    istruttori = [dict(row) for row in istruttori_raw]
    
    conn.close()
    
    return render_template('vista_istruttori.html', 
                         istruttori=istruttori,
                         anno_corrente=anno_corrente)

@app.route('/istruttore/<int:istruttore_id>')
@app.route('/istruttore/<int:istruttore_id>/<int:anno>')
def dettaglio_istruttore(istruttore_id, anno=2025):
    """Pagina dettaglio istruttore con statistiche e calendario"""
    conn = get_db()
    
    # Info istruttore
    istruttore = conn.execute('SELECT * FROM istruttori WHERE id = ?', (istruttore_id,)).fetchone()
    
    if not istruttore:
        conn.close()
        return "Istruttore non trovato", 404
    
    istruttore = dict(istruttore)
    
    # Impegni dell'anno
    impegni_raw = conn.execute('''
        SELECT i.*, ta.nome as attivita_nome, ta.colore, ta.categoria
        FROM impegni i
        JOIN tipi_attivita ta ON i.attivita_id = ta.id
        WHERE i.istruttore_id = ?
            AND (strftime('%Y', i.data_inizio) = ? OR strftime('%Y', i.data_fine) = ?)
        ORDER BY i.data_inizio
    ''', (istruttore_id, str(anno), str(anno))).fetchall()
    
    impegni = [dict(row) for row in impegni_raw]
    
    # Statistiche per categoria
    stats_categoria = {}
    for imp in impegni:
        cat = imp['categoria']
        if cat not in stats_categoria:
            stats_categoria[cat] = {'count': 0, 'giorni': 0}
        stats_categoria[cat]['count'] += 1
        stats_categoria[cat]['giorni'] += imp['giorni_lavorativi']
    
    # Sostituzioni OUT (quando questo istruttore è stato sostituito)
    sost_out_raw = conn.execute('''
        SELECT s.*, isost.nome as sostituto_nome, i.id_corso,
               ta.nome as attivita_nome, ta.colore
        FROM sostituzioni s
        JOIN istruttori isost ON s.istruttore_sostituto_id = isost.id
        JOIN impegni i ON s.impegno_id = i.id
        JOIN tipi_attivita ta ON i.attivita_id = ta.id
        WHERE s.istruttore_originale_id = ?
            AND strftime('%Y', s.data_sostituzione) = ?
        ORDER BY s.data_sostituzione
    ''', (istruttore_id, str(anno))).fetchall()
    
    sost_out = [dict(row) for row in sost_out_raw]
    
    # Sostituzioni IN (quando questo istruttore ha sostituito altri)
    sost_in_raw = conn.execute('''
        SELECT s.*, io.nome as originale_nome, i.id_corso,
               ta.nome as attivita_nome, ta.colore
        FROM sostituzioni s
        JOIN istruttori io ON s.istruttore_originale_id = io.id
        JOIN impegni i ON s.impegno_id = i.id
        JOIN tipi_attivita ta ON i.attivita_id = ta.id
        WHERE s.istruttore_sostituto_id = ?
            AND strftime('%Y', s.data_sostituzione) = ?
        ORDER BY s.data_sostituzione
    ''', (istruttore_id, str(anno))).fetchall()
    
    sost_in = [dict(row) for row in sost_in_raw]
    
    # Statistiche mensili (giorni per mese)
    from datetime import datetime
    stats_mensili = {i: 0 for i in range(1, 13)}  # 1-12 mesi
    
    for imp in impegni:
        data_inizio = datetime.strptime(imp['data_inizio'], '%Y-%m-%d')
        data_fine = datetime.strptime(imp['data_fine'], '%Y-%m-%d')
        
        # Per ogni giorno dell'impegno
        from datetime import timedelta
        current = data_inizio
        while current <= data_fine:
            if current.year == anno:
                # Solo giorni lavorativi (Lun-Ven)
                if current.weekday() < 5:
                    stats_mensili[current.month] += 1
            current += timedelta(days=1)
    
    # Totali
    totale_impegni = len(impegni)
    totale_giorni = sum(imp['giorni_lavorativi'] for imp in impegni)
    totale_giorni_netti = totale_giorni - len(sost_out) + len(sost_in)
    
    # Calcola giorni dell'anno per calendario
    num_giorni = 366 if cal_module.isleap(anno) else 365
    
    # Festivi italiani
    festivi = get_festivi_italiani(anno)
    
    # Crea struttura mesi
    mesi = []
    for mese in range(1, 13):
        giorni_mese = cal_module.monthrange(anno, mese)[1]
        mesi.append({
            'numero': mese,
            'nome': MESI_ITALIANI[mese],
            'giorni': giorni_mese
        })
    
    conn.close()
    
    return render_template('dettaglio_istruttore.html',
                         istruttore=istruttore,
                         anno=anno,
                         impegni=impegni,
                         stats_categoria=stats_categoria,
                         sost_out=sost_out,
                         sost_in=sost_in,
                         stats_mensili=stats_mensili,
                         totale_impegni=totale_impegni,
                         totale_giorni=totale_giorni,
                         totale_giorni_netti=totale_giorni_netti,
                         num_giorni=num_giorni,
                         mesi=mesi,
                         festivi=festivi)

@app.route('/vista_corsi')
def vista_corsi():
    """Pagina per selezionare un corso e vedere il suo calendario"""
    conn = get_db()
    
    # Lista di tutti i corsi (id_corso univoci) con info
    corsi_raw = conn.execute('''
        SELECT DISTINCT i.id_corso,
               MIN(i.data_inizio) as data_inizio,
               MAX(i.data_fine) as data_fine,
               COUNT(DISTINCT i.istruttore_id) as num_istruttori,
               COUNT(i.id) as num_impegni
        FROM impegni i
        WHERE i.id_corso IS NOT NULL AND i.id_corso != ''
        GROUP BY i.id_corso
        ORDER BY data_inizio DESC
    ''').fetchall()
    
    corsi = [dict(row) for row in corsi_raw]
    
    conn.close()
    
    return render_template('vista_corsi.html', corsi=corsi)

@app.route('/nuovo-corso')
def nuovo_corso():
    """Pagina wizard per creare un nuovo corso"""
    return render_template('nuovo_corso.html')
    
    corsi = [dict(row) for row in corsi_raw]
    
    conn.close()
    
    return render_template('vista_corsi.html', corsi=corsi)

@app.route('/istruttori')
def gestione_istruttori():
    """Pagina gestione istruttori"""
    conn = get_db()
    
    # Lista tutti gli istruttori (anche disattivati)
    istruttori = conn.execute('SELECT * FROM istruttori ORDER BY nome').fetchall()
    
    conn.close()
    
    return render_template('istruttori.html', istruttori=istruttori)

@app.route('/tipi-attivita')
def gestione_attivita():
    """Pagina gestione tipi attività"""
    conn = get_db()
    
    # Lista tutti i tipi attività
    attivita = conn.execute('SELECT * FROM tipi_attivita ORDER BY categoria, nome').fetchall()
    
    conn.close()
    
    return render_template('attivita.html', attivita=attivita)

@app.route('/report-giorni')
def report_giorni():
    """Pagina report giorni lavoro per istruttore"""
    conn = get_db()
    
    # Lista istruttori
    istruttori = conn.execute('SELECT * FROM istruttori WHERE attivo = 1 ORDER BY nome').fetchall()
    
    # Anni disponibili (dal primo impegno all'ultimo)
    anni_query = conn.execute('''
        SELECT DISTINCT strftime('%Y', data_inizio) as anno
        FROM impegni
        ORDER BY anno
    ''').fetchall()
    
    anni = [int(row['anno']) for row in anni_query] if anni_query else [2025]
    if not anni:
        anni = [2025, 2026, 2027, 2028, 2029, 2030]
    
    conn.close()
    
    return render_template('report_giorni.html', istruttori=istruttori, anni=anni)

@app.route('/festivi')
def gestione_festivi():
    """Pagina gestione festività personalizzate"""
    return render_template('festivi.html')

@app.route('/admin/utenti')
@require_role('Admin')
def admin_utenti():
    """Panel amministrazione utenti"""
    utenti = lista_utenti()
    conn = get_db()
    ruoli = conn.execute('SELECT * FROM ruoli').fetchall()
    conn.close()
    return render_template('admin_users.html', utenti=utenti, ruoli=ruoli)

@app.route('/admin/audit-log')
@require_role('Admin')
def admin_audit_log():
    """Visualizza audit log"""
    log = get_audit_log(500)
    return render_template('audit_log.html', log=log)

@app.route('/errore')
def errore():
    """Pagina errore generico"""
    messaggio = request.args.get('messaggio', 'Errore sconosciuto')
    return render_template('errore.html', messaggio=messaggio), 400

# ============================================================================
# API ENDPOINTS
# ============================================================================

@app.route('/api/impegni', methods=['GET'])
def api_get_impegni():
    """Ottieni lista impegni, opzionalmente filtrati per data"""
    conn = get_db()
    
    # Controlla se c'è un filtro per data
    data_filtro = request.args.get('data')
    
    if data_filtro:
        # Filtra impegni che includono questa data
        impegni = conn.execute('''
            SELECT i.*, ist.nome as istruttore_nome, ta.nome as attivita_nome, ta.colore
            FROM impegni i
            JOIN istruttori ist ON i.istruttore_id = ist.id
            JOIN tipi_attivita ta ON i.attivita_id = ta.id
            WHERE ? BETWEEN i.data_inizio AND i.data_fine
            ORDER BY i.data_inizio DESC
        ''', (data_filtro,)).fetchall()
    else:
        # Nessun filtro, restituisci tutti
        impegni = conn.execute('''
            SELECT i.*, ist.nome as istruttore_nome, ta.nome as attivita_nome, ta.colore
            FROM impegni i
            JOIN istruttori ist ON i.istruttore_id = ist.id
            JOIN tipi_attivita ta ON i.attivita_id = ta.id
            ORDER BY i.data_inizio DESC
        ''').fetchall()
    
    conn.close()
    
    return jsonify([dict(imp) for imp in impegni])

@app.route('/api/impegni', methods=['POST'])
@require_editor
def api_create_impegno():
    """Crea nuovo impegno"""
    try:
        data = request.json
        
        # Validazione campi obbligatori
        if not data.get('istruttore_id'):
            return jsonify({'error': 'Manca istruttore_id'}), 400
        if not data.get('attivita_id'):
            return jsonify({'error': 'Manca attivita_id'}), 400
        if not data.get('data_inizio'):
            return jsonify({'error': 'Manca data_inizio'}), 400
        
        data_fine_input = data.get('data_fine') or None
        giorni_input = data.get('giorni_lavorativi')

        if (giorni_input is None or giorni_input == '') and not data_fine_input:
            return jsonify({'error': 'Inserire giorni_lavorativi oppure data_fine'}), 400
        
        # Prepara giorni extra (solo se non vuoti)
        giorni_extra = []
        for i in range(1, 4):
            extra = data.get(f'giorno_extra_{i}')
            if extra and extra.strip():  # Controlla che non sia stringa vuota
                giorni_extra.append(extra)
        
        # Determina giorni_lavorativi e data_fine
        if data_fine_input:
            giorni_lavorativi = giorni_lavorativi_tra(data['data_inizio'], data_fine_input)
            data_fine = data_fine_input
        else:
            giorni_lavorativi = int(giorni_input)
            data_fine = calcola_data_fine(
                data['data_inizio'], 
                giorni_lavorativi,
                giorni_extra if giorni_extra else None
            )
        
        # VALIDAZIONE SOVRAPPOSIZIONI (se non è forzato)
        force = data.get('force', False)
        if not force:
            sovrapposizioni = verifica_sovrapposizione(
                data['istruttore_id'],
                data['data_inizio'],
                data_fine
            )
            
            if sovrapposizioni:
                dettagli = [f"{s['attivita_nome']} ({s['data_inizio']} - {s['data_fine']})" for s in sovrapposizioni]
                return jsonify({
                    'error': 'Sovrapposizione rilevata',
                    'conflitti': sovrapposizioni,
                    'messaggio': f"L'istruttore ha già impegni sovrapposti: {', '.join(dettagli)}"
                }), 409
        
        conn = get_db()
        c = conn.cursor()
        c.execute('''
            INSERT INTO impegni 
            (id_corso, istruttore_id, attivita_id, data_inizio, giorni_lavorativi, 
             giorno_extra_1, giorno_extra_2, giorno_extra_3, data_fine, note,
             luogo, aula, posti)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data.get('id_corso', ''),
            data['istruttore_id'],
            data['attivita_id'],
            data['data_inizio'],
            giorni_lavorativi,
            giorni_extra[0] if len(giorni_extra) > 0 else None,
            giorni_extra[1] if len(giorni_extra) > 1 else None,
            giorni_extra[2] if len(giorni_extra) > 2 else None,
            data_fine,
            data.get('note', ''),
            data.get('luogo'),
            data.get('aula'),
            data.get('posti')
        ))
        
        impegno_id = c.lastrowid
        conn.commit()
        conn.close()

        # Traccia creazione impegno
        log_action(
            "CREATE",
            f"Impegno {impegno_id} creato per istruttore {data['istruttore_id']} (attivita {data['attivita_id']}) dal {data['data_inizio']} per {giorni_lavorativi} giorni",
            request.headers.get('User-Agent', ''),
        )
        
        return jsonify({'success': True, 'id': impegno_id})
    except KeyError as e:
        return jsonify({'error': f'Campo mancante: {str(e)}'}), 400
    except Exception as e:
        return jsonify({'error': f'Errore interno: {str(e)}'}), 500

@app.route('/api/impegni/<int:impegno_id>', methods=['PUT'])
def api_update_impegno(impegno_id):
    """Aggiorna impegno esistente"""
    try:
        data = request.json
        
        # Validazione campi obbligatori
        if not data.get('istruttore_id'):
            return jsonify({'error': 'Manca istruttore_id'}), 400
        if not data.get('attivita_id'):
            return jsonify({'error': 'Manca attivita_id'}), 400
        if not data.get('data_inizio'):
            return jsonify({'error': 'Manca data_inizio'}), 400
        
        data_fine_input = data.get('data_fine') or None
        giorni_input = data.get('giorni_lavorativi')

        if (giorni_input is None or giorni_input == '') and not data_fine_input:
            return jsonify({'error': 'Inserire giorni_lavorativi oppure data_fine'}), 400
        
        # Prepara giorni extra (solo se non vuoti)
        giorni_extra = []
        for i in range(1, 4):
            extra = data.get(f'giorno_extra_{i}')
            if extra and extra.strip():  # Controlla che non sia stringa vuota
                giorni_extra.append(extra)
        
        # Determina giorni_lavorativi e data_fine
        if data_fine_input:
            giorni_lavorativi = giorni_lavorativi_tra(data['data_inizio'], data_fine_input)
            data_fine = data_fine_input
        else:
            giorni_lavorativi = int(giorni_input)
            data_fine = calcola_data_fine(
                data['data_inizio'], 
                giorni_lavorativi,
                giorni_extra if giorni_extra else None
            )
        
        # VALIDAZIONE SOVRAPPOSIZIONI (escluso impegno corrente, se non è forzato)
        force = data.get('force', False)
        if not force:
            sovrapposizioni = verifica_sovrapposizione(
                data['istruttore_id'],
                data['data_inizio'],
                data_fine,
                impegno_id_escluso=impegno_id
            )
            
            if sovrapposizioni:
                dettagli = [f"{s['attivita_nome']} ({s['data_inizio']} - {s['data_fine']})" for s in sovrapposizioni]
                return jsonify({
                    'error': 'Sovrapposizione rilevata',
                    'conflitti': sovrapposizioni,
                    'messaggio': f"L'istruttore ha già impegni sovrapposti: {', '.join(dettagli)}"
                }), 409
        
        conn = get_db()
        conn.execute('''
            UPDATE impegni 
            SET id_corso = ?, istruttore_id = ?, attivita_id = ?, 
                data_inizio = ?, giorni_lavorativi = ?, 
                giorno_extra_1 = ?, giorno_extra_2 = ?, giorno_extra_3 = ?,
                data_fine = ?, note = ?,
                luogo = ?, aula = ?, posti = ?,
                modificato_il = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (
            data.get('id_corso', ''),
            data['istruttore_id'],
            data['attivita_id'],
            data['data_inizio'],
            giorni_lavorativi,
            giorni_extra[0] if len(giorni_extra) > 0 else None,
            giorni_extra[1] if len(giorni_extra) > 1 else None,
            giorni_extra[2] if len(giorni_extra) > 2 else None,
            data_fine,
            data.get('note', ''),
            data.get('luogo'),
            data.get('aula'),
            data.get('posti'),
            impegno_id
        ))
        conn.commit()
        conn.close()

        # Traccia modifica impegno
        log_action(
            "UPDATE",
            f"Impegno {impegno_id} aggiornato (istruttore {data['istruttore_id']}, attivita {data['attivita_id']}, inizio {data['data_inizio']}, giorni {data['giorni_lavorativi']})",
            request.headers.get('User-Agent', ''),
        )
        
        return jsonify({'success': True})
    except KeyError as e:
        return jsonify({'error': f'Campo mancante: {str(e)}'}), 400
    except Exception as e:
        return jsonify({'error': f'Errore interno: {str(e)}'}), 500

@app.route('/api/impegni/<int:impegno_id>', methods=['DELETE'])
def api_delete_impegno(impegno_id):
    """Elimina impegno"""
    conn = get_db()
    conn.execute('DELETE FROM impegni WHERE id = ?', (impegno_id,))
    conn.commit()
    conn.close()

    # Traccia eliminazione impegno
    log_action(
        "DELETE",
        f"Impegno {impegno_id} eliminato",
        request.headers.get('User-Agent', ''),
    )
    
    return jsonify({'success': True})

@app.route('/api/impegni/<int:impegno_id>/shift', methods=['POST'])
def api_shift_impegno(impegno_id):
    """Sposta in avanti l'impegno di N giorni lavorativi e ricalcola data_fine.
    Body: {"days": int}
    """
    try:
        data = request.json or {}
        days = int(data.get('days', 0))
        if days <= 0:
            return jsonify({'error': 'days deve essere > 0'}), 400

        conn = get_db()
        imp = conn.execute('''
            SELECT id, istruttore_id, attivita_id, data_inizio, data_fine, giorni_lavorativi,
                   giorno_extra_1, giorno_extra_2, giorno_extra_3, id_corso, note
            FROM impegni WHERE id = ?
        ''', (impegno_id,)).fetchone()
        if not imp:
            conn.close()
            return jsonify({'error': 'Impegno non trovato'}), 404

        nuova_data_inizio = aggiungi_giorni_lavorativi(imp['data_inizio'], days)

        # Manteniamo i giorni_lavorativi e ricalcoliamo la data_fine con eventuali extra
        giorni_extra = [imp['giorno_extra_1'], imp['giorno_extra_2'], imp['giorno_extra_3']]
        nuova_data_fine = calcola_data_fine(nuova_data_inizio, int(imp['giorni_lavorativi']), giorni_extra)

        # Validazione sovrapposizioni (escludendo l'impegno stesso)
        conflitti = verifica_sovrapposizione(imp['istruttore_id'], nuova_data_inizio, nuova_data_fine, impegno_id_escluso=impegno_id)
        if conflitti:
            dettagli = [f"{c['attivita_nome']} ({c['data_inizio']} - {c['data_fine']})" for c in conflitti]
            conn.close()
            return jsonify({'error': 'Sovrapposizione rilevata', 'conflitti': conflitti, 'messaggio': ', '.join(dettagli)}), 409

        conn.execute('''
            UPDATE impegni SET data_inizio = ?, data_fine = ?, modificato_il = CURRENT_TIMESTAMP WHERE id = ?
        ''', (nuova_data_inizio, nuova_data_fine, impegno_id))
        conn.commit()
        conn.close()

        log_action(
            "UPDATE",
            f"Impegno {impegno_id} spostato di {days} giorni lavorativi (nuovo inizio {nuova_data_inizio}, fine {nuova_data_fine})",
            request.headers.get('User-Agent', ''),
        )

        return jsonify({'success': True, 'data_inizio': nuova_data_inizio, 'data_fine': nuova_data_fine})
    except Exception as e:
        return jsonify({'error': f'Errore interno: {str(e)}'}), 500

@app.route('/api/calendario/<int:anno>')
def api_calendario(anno):
    """API per ottenere dati calendario"""
    conn = get_db()
    
    # Impegni dell'anno con dettagli
    impegni = conn.execute('''
        SELECT i.*, ist.nome as istruttore_nome, ta.nome as attivita_nome, ta.colore
        FROM impegni i
        JOIN istruttori ist ON i.istruttore_id = ist.id
        JOIN tipi_attivita ta ON i.attivita_id = ta.id
        WHERE 
            (strftime('%Y', i.data_inizio) = ? OR strftime('%Y', i.data_fine) = ?)
            OR (i.data_inizio <= ? AND i.data_fine >= ?)
    ''', (str(anno), str(anno), f"{anno}-12-31", f"{anno}-01-01")).fetchall()
    
    conn.close()
    
    return jsonify([dict(imp) for imp in impegni])

@app.route('/api/sovrapposizioni')
def api_sovrapposizioni():
    """Trova sovrapposizioni tra impegni dello stesso istruttore, escludendo giorni con sostituzioni"""
    conn = get_db()
    
    # Trova tutte le sovrapposizioni
    sovrapposizioni_raw = conn.execute('''
        SELECT 
            i1.id as id1, i2.id as id2,
            i1.istruttore_id as istruttore_id,
            ist.nome as istruttore,
            ta1.nome as attivita1, i1.data_inizio as inizio1, i1.data_fine as fine1,
            ta2.nome as attivita2, i2.data_inizio as inizio2, i2.data_fine as fine2,
            ta1.categoria as categoria1, ta2.categoria as categoria2
        FROM impegni i1
        JOIN impegni i2 ON i1.istruttore_id = i2.istruttore_id AND i1.id < i2.id
        JOIN istruttori ist ON i1.istruttore_id = ist.id
        JOIN tipi_attivita ta1 ON i1.attivita_id = ta1.id
        JOIN tipi_attivita ta2 ON i2.attivita_id = ta2.id
        WHERE 
            i1.data_inizio <= i2.data_fine 
            AND i1.data_fine >= i2.data_inizio
        ORDER BY ist.nome, i1.data_inizio
    ''').fetchall()
    
    # Recupera tutte le sostituzioni per filtrare i conflitti risolti
    sostituzioni_set = set()
    sostituzioni_raw = conn.execute('''
        SELECT impegno_id, istruttore_originale_id, data_sostituzione
        FROM sostituzioni
    ''').fetchall()
    for s in sostituzioni_raw:
        sostituzioni_set.add((s['impegno_id'], s['istruttore_originale_id'], s['data_sostituzione']))
    
    conn.close()
    
    # Filtra sovrapposizioni: escludi quelle completamente risolte con sostituzioni
    from datetime import datetime
    sovrapposizioni_filtrate = []
    for sov in sovrapposizioni_raw:
        # Calcola range di sovrapposizione
        start = max(sov['inizio1'], sov['inizio2'])
        end = min(sov['fine1'], sov['fine2'])
        
        # Verifica se tutti i giorni sovrapposti hanno sostituzioni
        d1 = datetime.strptime(start, '%Y-%m-%d')
        d2 = datetime.strptime(end, '%Y-%m-%d')
        giorni_conflitto = []
        while d1 <= d2:
            data_str = d1.strftime('%Y-%m-%d')
            # Controlla se questo giorno ha una sostituzione per uno dei due impegni
            ha_sost_i1 = (sov['id1'], sov['istruttore_id'], data_str) in sostituzioni_set
            ha_sost_i2 = (sov['id2'], sov['istruttore_id'], data_str) in sostituzioni_set
            if not ha_sost_i1 and not ha_sost_i2:
                giorni_conflitto.append(data_str)
            d1 += timedelta(days=1)
        
        # Aggiungi solo se ci sono giorni senza sostituzione
        if giorni_conflitto:
            sovrapposizioni_filtrate.append(dict(sov))
    
    return jsonify(sovrapposizioni_filtrate)

@app.route('/api/statistiche')
def api_statistiche():
    """Statistiche generali"""
    conn = get_db()
    
    stats = {
        'totale_impegni': conn.execute('SELECT COUNT(*) as cnt FROM impegni').fetchone()['cnt'],
        'totale_corsi': conn.execute('SELECT COUNT(DISTINCT id_corso) as cnt FROM impegni WHERE id_corso != ""').fetchone()['cnt'],
        'totale_istruttori': conn.execute('SELECT COUNT(*) as cnt FROM istruttori WHERE attivo = 1').fetchone()['cnt'],
        
        # Impegni per istruttore
        'per_istruttore': [dict(row) for row in conn.execute('''
            SELECT ist.nome, COUNT(*) as totale
            FROM impegni i
            JOIN istruttori ist ON i.istruttore_id = ist.id
            GROUP BY ist.nome
            ORDER BY totale DESC
        ''').fetchall()],
        
        # Impegni per tipo
        'per_tipo': [dict(row) for row in conn.execute('''
            SELECT ta.nome, ta.categoria, COUNT(*) as totale
            FROM impegni i
            JOIN tipi_attivita ta ON i.attivita_id = ta.id
            GROUP BY ta.nome
            ORDER BY totale DESC
        ''').fetchall()],
    }
    
    conn.close()
    
    return jsonify(stats)

@app.route('/api/dashboard')
def api_dashboard():
    """Dashboard giornaliera - impegni di oggi"""
    from datetime import date
    conn = get_db()
    
    oggi = date.today().strftime('%Y-%m-%d')
    
    # Impegni attivi oggi
    impegni_oggi = conn.execute("""
        SELECT i.*, 
               ist.nome as istruttore_nome,
               ta.nome as attivita_nome,
               ta.colore,
               ta.categoria
        FROM impegni i
        JOIN istruttori ist ON i.istruttore_id = ist.id
        JOIN tipi_attivita ta ON i.attivita_id = ta.id
        WHERE ? BETWEEN i.data_inizio AND i.data_fine
        ORDER BY ist.nome, i.data_inizio
    """, (oggi,)).fetchall()
    
    # Statistiche giornaliere
    stats_oggi = {
        'totale_impegni': len(impegni_oggi),
        'corsi_attivi': len(set(imp['id_corso'] for imp in impegni_oggi if imp['id_corso'])),
        'istruttori_occupati': len(set(imp['istruttore_id'] for imp in impegni_oggi))
    }
    
    # Prossimi 7 giorni - impegni per data
    from datetime import timedelta
    prossimi_giorni = []
    for i in range(7):
        data = (date.today() + timedelta(days=i)).strftime('%Y-%m-%d')
        count = conn.execute("""
            SELECT COUNT(*) as cnt 
            FROM impegni 
            WHERE ? BETWEEN data_inizio AND data_fine
        """, (data,)).fetchone()['cnt']
        prossimi_giorni.append({
            'data': data,
            'count': count
        })
    
    conn.close()
    
    return jsonify({
        'oggi': oggi,
        'impegni': [dict(row) for row in impegni_oggi],
        'stats': stats_oggi,
        'prossimi_7_giorni': prossimi_giorni
    })


# ============================================================================
# API GESTIONE ISTRUTTORI
# ============================================================================

@app.route('/api/istruttori', methods=['GET'])
def api_get_istruttori():
    """Ottieni lista istruttori"""
    conn = get_db()
    istruttori = conn.execute('SELECT * FROM istruttori ORDER BY nome').fetchall()
    conn.close()
    return jsonify([dict(i) for i in istruttori])

@app.route('/api/istruttori', methods=['POST'])
def api_create_istruttore():
    """Crea nuovo istruttore"""
    data = request.json
    
    conn = get_db()
    c = conn.cursor()
    
    try:
        c.execute('''
            INSERT INTO istruttori (nome, email, attivo)
            VALUES (?, ?, 1)
        ''', (data['nome'], data.get('email', '')))
        
        istruttore_id = c.lastrowid
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'id': istruttore_id})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/istruttori/<int:istruttore_id>', methods=['PUT'])
def api_update_istruttore(istruttore_id):
    """Aggiorna istruttore"""
    data = request.json
    
    conn = get_db()
    conn.execute('''
        UPDATE istruttori 
        SET nome = ?, email = ?, attivo = ?
        WHERE id = ?
    ''', (data['nome'], data.get('email', ''), data.get('attivo', 1), istruttore_id))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})

@app.route('/api/istruttori/<int:istruttore_id>/disattiva', methods=['POST'])
def api_disattiva_istruttore(istruttore_id):
    """Disattiva istruttore (pensione)"""
    conn = get_db()
    conn.execute('UPDATE istruttori SET attivo = 0 WHERE id = ?', (istruttore_id,))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})

@app.route('/api/istruttori/<int:istruttore_id>/attiva', methods=['POST'])
def api_attiva_istruttore(istruttore_id):
    """Riattiva istruttore"""
    conn = get_db()
    conn.execute('UPDATE istruttori SET attivo = 1 WHERE id = ?', (istruttore_id,))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})

# ============================================================================
# API GESTIONE TIPI ATTIVITÀ
# ============================================================================

@app.route('/api/tipi-attivita', methods=['GET'])
def api_get_tipi_attivita():
    """Ottieni lista tipi attività"""
    conn = get_db()
    attivita = conn.execute('SELECT * FROM tipi_attivita ORDER BY categoria, nome').fetchall()
    conn.close()
    return jsonify([dict(a) for a in attivita])

@app.route('/api/tipi-attivita', methods=['POST'])
def api_create_tipo_attivita():
    """Crea nuovo tipo attività"""
    data = request.json
    
    conn = get_db()
    c = conn.cursor()
    
    try:
        c.execute('''
            INSERT INTO tipi_attivita (nome, colore, categoria)
            VALUES (?, ?, ?)
        ''', (data['nome'], data['colore'], data.get('categoria', 'ALTRO')))
        
        attivita_id = c.lastrowid
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'id': attivita_id})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/tipi-attivita/<int:attivita_id>', methods=['PUT'])
def api_update_tipo_attivita(attivita_id):
    """Aggiorna tipo attività"""
    data = request.json
    
    conn = get_db()
    conn.execute('''
        UPDATE tipi_attivita 
        SET nome = ?, colore = ?, categoria = ?
        WHERE id = ?
    ''', (data['nome'], data['colore'], data.get('categoria', 'ALTRO'), attivita_id))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})

@app.route('/api/tipi-attivita/<int:attivita_id>', methods=['DELETE'])
def api_delete_tipo_attivita(attivita_id):
    """Elimina tipo attività (solo se non usato)"""
    conn = get_db()
    
    # Verifica se è usato
    count = conn.execute('SELECT COUNT(*) as cnt FROM impegni WHERE attivita_id = ?', (attivita_id,)).fetchone()['cnt']
    
    if count > 0:
        conn.close()
        return jsonify({'success': False, 'error': f'Tipo attività usato in {count} impegni'}), 400
    
    conn.execute('DELETE FROM tipi_attivita WHERE id = ?', (attivita_id,))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})

# ============================================================================
# API SOSTITUZIONI
# ============================================================================

@app.route('/api/sostituzioni', methods=['GET'])
def api_get_sostituzioni():
    """Ottieni lista sostituzioni"""
    impegno_id = request.args.get('impegno_id')
    
    conn = get_db()
    
    if impegno_id:
        sostituzioni = conn.execute('''
            SELECT s.*, 
                   io.nome as istruttore_originale, 
                   isost.nome as istruttore_sostituto
            FROM sostituzioni s
            JOIN istruttori io ON s.istruttore_originale_id = io.id
            JOIN istruttori isost ON s.istruttore_sostituto_id = isost.id
            WHERE s.impegno_id = ?
            ORDER BY s.data_sostituzione
        ''', (impegno_id,)).fetchall()
    else:
        sostituzioni = conn.execute('''
            SELECT s.*, 
                   io.nome as istruttore_originale, 
                   isost.nome as istruttore_sostituto,
                   i.id_corso, ta.nome as attivita_nome
            FROM sostituzioni s
            JOIN istruttori io ON s.istruttore_originale_id = io.id
            JOIN istruttori isost ON s.istruttore_sostituto_id = isost.id
            JOIN impegni i ON s.impegno_id = i.id
            JOIN tipi_attivita ta ON i.attivita_id = ta.id
            ORDER BY s.data_sostituzione DESC
        ''').fetchall()
    
    conn.close()
    return jsonify([dict(s) for s in sostituzioni])

@app.route('/api/sostituzioni', methods=['POST'])
def api_create_sostituzione():
    """Crea nuova sostituzione"""
    data = request.json
    
    conn = get_db()
    c = conn.cursor()
    
    try:
        # Verifica che la data sia nell'intervallo dell'impegno
        impegno = conn.execute('''
            SELECT data_inizio, data_fine, istruttore_id
            FROM impegni WHERE id = ?
        ''', (data['impegno_id'],)).fetchone()
        
        if not impegno:
            return jsonify({'success': False, 'error': 'Impegno non trovato'}), 404
        
        data_sost = data['data_sostituzione']
        if not (impegno['data_inizio'] <= data_sost <= impegno['data_fine']):
            return jsonify({'success': False, 'error': 'Data non nell\'intervallo dell\'impegno'}), 400
        
        # VERIFICA DISPONIBILITÀ SOSTITUTO
        # Controlla se il sostituto ha già impegni in quella data
        conflitti_sostituto = conn.execute('''
            SELECT i.*, ta.nome as attivita_nome
            FROM impegni i
            JOIN tipi_attivita ta ON i.attivita_id = ta.id
            WHERE i.istruttore_id = ?
            AND i.data_inizio <= ?
            AND i.data_fine >= ?
        ''', (data['istruttore_sostituto_id'], data_sost, data_sost)).fetchall()
        
        if conflitti_sostituto:
            dettagli = [f"{c['attivita_nome']} ({c['data_inizio']} - {c['data_fine']})" for c in conflitti_sostituto]
            conn.close()
            messaggio = f"L'istruttore sostituto ha già impegni: {', '.join(dettagli)}"
            return jsonify({
                'success': False,
                'error': 'Sostituto non disponibile',
                'messaggio': messaggio,
                'message': messaggio
            }), 409
        
        c.execute('''
            INSERT INTO sostituzioni 
            (impegno_id, data_sostituzione, istruttore_originale_id, istruttore_sostituto_id, note)
            VALUES (?, ?, ?, ?, ?)
        ''', (
            data['impegno_id'],
            data['data_sostituzione'],
            impegno['istruttore_id'],
            data['istruttore_sostituto_id'],
            data.get('note', '')
        ))
        
        sostituzione_id = c.lastrowid
        conn.commit()
        conn.close()

        # Traccia creazione sostituzione
        log_action(
            "CREATE",
            f"Sostituzione {sostituzione_id} su impegno {data['impegno_id']} (orig {impegno['istruttore_id']} -> sost {data['istruttore_sostituto_id']}) in data {data['data_sostituzione']}",
            request.headers.get('User-Agent', ''),
        )
        
        return jsonify({'success': True, 'id': sostituzione_id})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/sostituzioni/<int:sostituzione_id>', methods=['DELETE'])
def api_delete_sostituzione(sostituzione_id):
    """Elimina sostituzione"""
    conn = get_db()
    conn.execute('DELETE FROM sostituzioni WHERE id = ?', (sostituzione_id,))
    conn.commit()
    conn.close()

    # Traccia eliminazione sostituzione
    log_action(
        "DELETE",
        f"Sostituzione {sostituzione_id} eliminata",
        request.headers.get('User-Agent', ''),
    )
    
    return jsonify({'success': True})

# ============================================================================
# API FESTIVI PERSONALIZZATI
# ============================================================================

@app.route('/api/festivi-custom', methods=['GET'])
def api_get_festivi_custom():
    """Ottieni lista festivi personalizzati"""
    anno = request.args.get('anno')
    
    conn = get_db()
    
    if anno:
        festivi = conn.execute(
            'SELECT * FROM festivi_custom WHERE strftime("%Y", data) = ? ORDER BY data',
            (anno,)
        ).fetchall()
    else:
        festivi = conn.execute('SELECT * FROM festivi_custom ORDER BY data').fetchall()
    
    conn.close()
    return jsonify([dict(f) for f in festivi])

@app.route('/api/festivi-custom', methods=['POST'])
def api_create_festivo_custom():
    """Crea nuovo festivo personalizzato"""
    data = request.json
    
    conn = get_db()
    try:
        c = conn.cursor()
        c.execute('''
            INSERT INTO festivi_custom (data, descrizione, tipo)
            VALUES (?, ?, ?)
        ''', (data['data'], data['descrizione'], data.get('tipo', 'AZIENDALE')))
        
        festivo_id = c.lastrowid
        conn.commit()
        
        log_action("CREATE", f"Festivo custom {festivo_id} creato: {data['data']} - {data['descrizione']}", 
                   request.headers.get('User-Agent', ''))
        
        return jsonify({'success': True, 'id': festivo_id})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()  # Assicura che si chiude sempre

@app.route('/api/festivi-custom/<int:festivo_id>', methods=['DELETE'])
def api_delete_festivo_custom(festivo_id):
    """Elimina festivo personalizzato"""
    conn = get_db()
    conn.execute('DELETE FROM festivi_custom WHERE id = ?', (festivo_id,))
    conn.commit()
    conn.close()
    
    log_action("DELETE", f"Festivo custom {festivo_id} eliminato", 
               request.headers.get('User-Agent', ''))
    
    return jsonify({'success': True})

# ============================================================================
# API GESTIONE CORSI
# ============================================================================

@app.route('/api/calcola-data-fine', methods=['POST'])
def api_calcola_data_fine():
    """Calcola data fine dato data inizio e giorni lavorativi"""
    try:
        data = request.json
        data_inizio = data.get('data_inizio')
        giorni_lavorativi = int(data.get('giorni_lavorativi', 0))
        
        if not data_inizio or giorni_lavorativi <= 0:
            return jsonify({'error': 'Parametri mancanti'}), 400
        
        data_fine = calcola_data_fine(data_inizio, giorni_lavorativi)
        
        return jsonify({
            'success': True,
            'data_fine': data_fine
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/corsi', methods=['POST'])
@require_editor
def api_create_corso():
    """Crea un nuovo corso con impegni associati"""
    try:
        data = request.json
        
        # Validazione
        if not data.get('id_corso'):
            return jsonify({'error': 'Manca id_corso'}), 400
        if not data.get('data_inizio') or not data.get('data_fine'):
            return jsonify({'error': 'Manca periodo corso'}), 400
        if not data.get('istruttori') or len(data['istruttori']) == 0:
            return jsonify({'error': 'Nessun istruttore selezionato'}), 400
        if not data.get('attivita') or len(data['attivita']) == 0:
            return jsonify({'error': 'Nessuna attività aggiunta'}), 400
        
        conn = get_db()
        c = conn.cursor()
        
        # Crea impegni per ogni attività
        impegni_creati = []
        for att in data['attivita']:
            # Valida istruttore
            ist = conn.execute('SELECT id FROM istruttori WHERE id = ?', 
                              (att['istruttore_id'],)).fetchone()
            if not ist:
                conn.close()
                return jsonify({'error': f"Istruttore {att['istruttore_id']} non trovato"}), 400
            
            # Valida attività
            attivita = conn.execute(
                'SELECT id FROM tipi_attivita WHERE nome = ?',
                (att['attivita_nome'],)
            ).fetchone()
            
            if not attivita:
                conn.close()
                return jsonify({'error': f"Attività '{att['attivita_nome']}' non trovata"}), 400
            
            # Colleziona giorni extra
            giorni_extra = []
            if att.get('giorno_extra_1'):
                giorni_extra.append(att['giorno_extra_1'])
            if att.get('giorno_extra_2'):
                giorni_extra.append(att['giorno_extra_2'])
            if att.get('giorno_extra_3'):
                giorni_extra.append(att['giorno_extra_3'])
            
            # Calcola data fine considerando giorni extra
            data_fine = calcola_data_fine(att['data_inizio'], att['giorni_lavorativi'], giorni_extra if giorni_extra else None)
            
            # Verifica sovrapposizioni
            sovrapposizioni = verifica_sovrapposizione(
                att['istruttore_id'],
                att['data_inizio'],
                data_fine
            )
            
            if sovrapposizioni:
                conn.close()
                return jsonify({
                    'error': 'Sovrapposizione rilevata',
                    'messaggio': f"L'istruttore ha già impegni sovrapposti"
                }), 409
            
            # Crea impegno
            # Determina istruttore di riferimento per questo impegno
            istruttore_rif_id = data.get('istruttore_riferimento') if att['istruttore_id'] == data.get('istruttore_riferimento') else None
            
            c.execute('''
                INSERT INTO impegni
                (id_corso, istruttore_id, attivita_id, data_inizio, giorni_lavorativi,
                 data_fine, note, luogo, aula, posti, istruttore_riferimento,
                 giorno_extra_1, giorno_extra_2, giorno_extra_3)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                data['id_corso'],
                att['istruttore_id'],
                attivita['id'],
                att['data_inizio'],
                att['giorni_lavorativi'],
                data_fine,
                att.get('note', ''),
                data.get('luogo', ''),
                data.get('aula', ''),
                data.get('posti', ''),
                istruttore_rif_id,
                att.get('giorno_extra_1', ''),
                att.get('giorno_extra_2', ''),
                att.get('giorno_extra_3', '')
            ))
            
            impegni_creati.append(c.lastrowid)
        
        conn.commit()
        conn.close()
        
        # Traccia creazione corso
        log_action(
            "CREATE",
            f"Corso {data['id_corso']} creato con {len(impegni_creati)} impegni",
            request.headers.get('User-Agent', '')
        )
        
        return jsonify({
            'success': True,
            'id_corso': data['id_corso'],
            'impegni_creati': len(impegni_creati)
        })
    
    except Exception as e:
        return jsonify({'error': f'Errore interno: {str(e)}'}), 500

# ============================================================================
# API REPORT GIORNI LAVORO
# ============================================================================

@app.route('/api/report-giorni')
def api_report_giorni():
    """Calcola giorni lavoro per istruttore e anno"""
    istruttore_id = request.args.get('istruttore_id')
    anno = request.args.get('anno')
    
    conn = get_db()
    
    # Query per ottenere impegni dell'istruttore nell'anno
    impegni = conn.execute('''
        SELECT i.*, ta.nome as attivita_nome, ta.categoria
        FROM impegni i
        JOIN tipi_attivita ta ON i.attivita_id = ta.id
        WHERE i.istruttore_id = ?
        AND (
            strftime('%Y', i.data_inizio) = ?
            OR strftime('%Y', i.data_fine) = ?
            OR (i.data_inizio <= ? AND i.data_fine >= ?)
        )
    ''', (istruttore_id, anno, anno, f"{anno}-12-31", f"{anno}-01-01")).fetchall()
    
    # Sostituzioni dove questo istruttore è stato SOSTITUITO (da sottrarre)
    sostituzioni_out = conn.execute('''
        SELECT COUNT(*) as cnt
        FROM sostituzioni
        WHERE istruttore_originale_id = ?
        AND strftime('%Y', data_sostituzione) = ?
    ''', (istruttore_id, anno)).fetchone()['cnt']
    
    # Sostituzioni dove questo istruttore è SOSTITUTO (da aggiungere)
    sostituzioni_in = conn.execute('''
        SELECT COUNT(*) as cnt
        FROM sostituzioni
        WHERE istruttore_sostituto_id = ?
        AND strftime('%Y', data_sostituzione) = ?
    ''', (istruttore_id, anno)).fetchone()['cnt']
    
    conn.close()
    
    # Calcola totali per categoria
    totali = {
        'CORSO': 0,
        'ESAME': 0,
        'TIROCINIO': 0,
        'ADDESTRAMENTO': 0,
        'PRATICHE': 0,
        'ALTRO': 0,
        'ASSENZA': 0
    }
    
    for imp in impegni:
        categoria = imp['categoria'] if imp['categoria'] in totali else 'ALTRO'
        
        # Conta solo giorni nell'anno specificato
        data_inizio = datetime.strptime(imp['data_inizio'], '%Y-%m-%d')
        data_fine = datetime.strptime(imp['data_fine'], '%Y-%m-%d')
        
        inizio_anno = datetime(int(anno), 1, 1)
        fine_anno = datetime(int(anno), 12, 31)
        
        # Limita alle date dell'anno
        inizio_eff = max(data_inizio, inizio_anno)
        fine_eff = min(data_fine, fine_anno)
        
        # Calcola giorni lavorativi effettivi nell'intervallo
        giorni = imp['giorni_lavorativi']
        
        # Approssimazione: distribuzione proporzionale
        if data_fine > data_inizio:
            giorni_totali = (data_fine - data_inizio).days + 1
            giorni_anno = (fine_eff - inizio_eff).days + 1
            giorni_effettivi = int(giorni * giorni_anno / giorni_totali)
        else:
            giorni_effettivi = giorni
        
        totali[categoria] += giorni_effettivi
    
    # Applica sostituzioni
    totale_giorni = sum(totali.values())
    totale_giorni = totale_giorni - sostituzioni_out + sostituzioni_in
    
    return jsonify({
        'totali_categoria': totali,
        'totale_giorni': totale_giorni,
        'sostituzioni_out': sostituzioni_out,
        'sostituzioni_in': sostituzioni_in,
        'dettagli_impegni': len(impegni)
    })

# ============================================================================
# EXPORT EXCEL
# ============================================================================

@app.route('/export/excel')
def export_excel():
    """Genera ed esporta file Excel completo"""
    conn = get_db()
    
    # Ottieni tutti i dati
    istruttori = conn.execute('SELECT * FROM istruttori WHERE attivo = 1 ORDER BY nome').fetchall()
    attivita_list = conn.execute('SELECT * FROM tipi_attivita').fetchall()
    impegni_list = conn.execute('''
        SELECT i.*, ist.nome as istruttore_nome, ta.nome as attivita_nome, ta.colore
        FROM impegni i
        JOIN istruttori ist ON i.istruttore_id = ist.id
        JOIN tipi_attivita ta ON i.attivita_id = ta.id
        ORDER BY i.data_inizio
    ''').fetchall()
    
    conn.close()
    
    # Crea workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # DATABASE
    ws_db = wb.create_sheet("📊 DATABASE", 0)
    ws_db['A1'] = "📊 DATABASE IMPEGNI"
    ws_db['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws_db['A1'].fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    ws_db['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_db.merge_cells('A1:H1')
    ws_db.row_dimensions[1].height = 30
    
    # Intestazioni
    headers = ["ID", "ID CORSO", "ISTRUTTORE", "ATTIVITÀ", "DATA INIZIO", "GIORNI LAV.", "DATA FINE", "NOTE"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_db.cell(2, col_idx, header)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Dati
    for row_idx, imp in enumerate(impegni_list, 3):
        ws_db.cell(row_idx, 1, imp['id'])
        ws_db.cell(row_idx, 2, imp['id_corso'] or '')
        ws_db.cell(row_idx, 3, imp['istruttore_nome'])
        ws_db.cell(row_idx, 4, imp['attivita_nome'])
        ws_db.cell(row_idx, 5, imp['data_inizio'])
        ws_db.cell(row_idx, 6, imp['giorni_lavorativi'])
        ws_db.cell(row_idx, 7, imp['data_fine'])
        ws_db.cell(row_idx, 8, imp['note'] or '')
    
    # Salva in memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Calendario_Istruttori_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

# ============================================================================
# AUDIT LOG
# ============================================================================

@app.route('/audit-log')
def audit_log():
    """Visualizza il log di audit di tutte le azioni"""
    try:
        conn = get_db()
        logs = conn.execute('''
            SELECT * FROM audit_log
            ORDER BY timestamp DESC
            LIMIT 500
        ''').fetchall()
        conn.close()
        return render_template('audit_log.html', logs=logs)
    except Exception as e:
        # Mostra errore utile in pagina per debug
        try:
            return render_template('audit_log.html', logs=[], error=str(e))
        except Exception:
            return jsonify({
                'error': 'Errore caricamento audit log',
                'detail': str(e)
            }), 500

# ============================================================================
# API ADMIN - GESTIONE UTENTI
# ============================================================================

@app.route('/api/utenti', methods=['GET'])
@require_role('Admin')
def api_lista_utenti():
    """Lista tutti gli utenti"""
    utenti = lista_utenti()
    return jsonify(utenti)

@app.route('/api/utenti', methods=['POST'])
@require_role('Admin')
def api_crea_utente():
    """Crea nuovo utente"""
    try:
        data = request.json
        username = data.get('username', '').strip()
        email = data.get('email', '').strip()
        nome = data.get('nome', '').strip()
        cognome = data.get('cognome', '').strip()
        password = data.get('password', '')
        ruolo = data.get('ruolo', 'Viewer')
        
        if not username or not password or not ruolo:
            return jsonify({'errore': 'Campi obbligatori mancanti'}), 400
        
        result = crea_utente(username, email, nome, cognome, password, ruolo)
        if 'errore' in result:
            return jsonify(result), 400
        
        registra_audit(session.get('utente_id'), 'CREA_UTENTE', 'utenti', result['utente_id'],
                      f'Username: {username}, Ruolo: {ruolo}', request.remote_addr)
        
        return jsonify({'successo': True, 'utente_id': result['utente_id']})
    except Exception as e:
        return jsonify({'errore': str(e)}), 500

@app.route('/api/utenti/<int:utente_id>/disattiva', methods=['POST'])
@require_role('Admin')
def api_disattiva_utente(utente_id):
    """Disattiva utente"""
    try:
        if utente_id == session.get('utente_id'):
            return jsonify({'errore': 'Non puoi disattivare il tuo account'}), 400
        
        conn = get_db()
        conn.execute('UPDATE utenti SET attivo = 0 WHERE id = ?', (utente_id,))
        conn.commit()
        conn.close()
        
        registra_audit(session.get('utente_id'), 'DISATTIVA_UTENTE', 'utenti', utente_id,
                      f'Utente disattivato', request.remote_addr)
        
        return jsonify({'successo': True})
    except Exception as e:
        return jsonify({'errore': str(e)}), 500

@app.route('/api/utenti/<int:utente_id>/attiva', methods=['POST'])
@require_role('Admin')
def api_attiva_utente(utente_id):
    """Attiva utente"""
    try:
        conn = get_db()
        conn.execute('UPDATE utenti SET attivo = 1 WHERE id = ?', (utente_id,))
        conn.commit()
        conn.close()
        
        registra_audit(session.get('utente_id'), 'ATTIVA_UTENTE', 'utenti', utente_id,
                      f'Utente attivato', request.remote_addr)
        
        return jsonify({'successo': True})
    except Exception as e:
        return jsonify({'errore': str(e)}), 500

@app.route('/api/corsi/<path:id_corso>', methods=['GET'])
@require_login
def api_get_corso(id_corso):
    """Ottieni impegni di un corso"""
    conn = get_db()
    impegni = conn.execute('''
        SELECT i.*, ist.nome as istruttore_nome, ta.nome as attivita_nome, ta.colore
        FROM impegni i
        JOIN istruttori ist ON i.istruttore_id = ist.id
        JOIN tipi_attivita ta ON i.attivita_id = ta.id
        WHERE i.id_corso = ?
        ORDER BY i.data_inizio
    ''', (id_corso,)).fetchall()
    conn.close()
    
    return jsonify([dict(imp) for imp in impegni])

@app.route('/api/corsi/aggiorna', methods=['POST'])
@require_login
def api_aggiorna_corso():
    """Aggiorna i dettagli di un corso (date, luogo, aula, posti)"""
    try:
        data = request.json
        
        id_corso = data.get('id_corso')
        if not id_corso:
            return jsonify({'error': 'Manca id_corso'}), 400
        
        data_inizio = data.get('data_inizio')
        data_fine = data.get('data_fine')
        
        if not data_inizio or not data_fine:
            return jsonify({'error': 'Manca data_inizio o data_fine'}), 400
        
        # Validazione date
        if data_inizio > data_fine:
            return jsonify({'error': 'Data inizio deve essere prima di data fine'}), 400
        
        conn = get_db()
        c = conn.cursor()
        
        # Ottieni tutti gli impegni del corso
        impegni = c.execute('''
            SELECT id, data_inizio, data_fine, giorni_lavorativi
            FROM impegni
            WHERE id_corso = ?
        ''', (id_corso,)).fetchall()
        
        if not impegni:
            conn.close()
            return jsonify({'error': 'Nessun impegno trovato per questo corso'}), 404
        
        # Calcola giorni_lavorativi dalle nuove date
        giorni_totali = giorni_lavorativi_tra(data_inizio, data_fine)
        
        # Aggiorna tutti gli impegni del corso
        c.execute('''
            UPDATE impegni 
            SET data_inizio = ?, data_fine = ?, giorni_lavorativi = ?,
                luogo = ?, aula = ?, posti = ?,
                modificato_il = CURRENT_TIMESTAMP
            WHERE id_corso = ?
        ''', (
            data_inizio,
            data_fine,
            giorni_totali if giorni_totali > 0 else 1,  # Minimo 1 giorno
            data.get('luogo'),
            data.get('aula'),
            data.get('posti'),
            id_corso
        ))
        
        conn.commit()
        
        # Log audit
        try:
            log_audit(f"Aggiornato corso {id_corso} (date {data_inizio}-{data_fine}, {giorni_totali} giorni)")
        except:
            pass  # Ignora errori di logging
        
        conn.close()
        
        return jsonify({'success': True}), 200
    except Exception as e:
        print(f"Errore aggiorna_corso: {str(e)}")  # Log per debug
        return jsonify({'error': str(e)}), 500

@app.route('/api/audit-log', methods=['GET'])
@require_role('Admin')
def api_get_audit_log():
    """Ottieni audit log"""
    limit = request.args.get('limit', 500, type=int)
    log = get_audit_log(limit)
    return jsonify(log)

# ============================================================================
# MAIN
# ============================================================================

if __name__ == '__main__':
    print("🚀 Avvio Calendario Istruttori...")
    print("📍 Accedi a: http://localhost:5000")
    print("⚠️  Premi CTRL+C per fermare il server")
    print("")
    app.run(debug=True, host='0.0.0.0', port=5000)

